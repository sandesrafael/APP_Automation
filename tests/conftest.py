# -*- coding: utf-8 -*-
import os
os.environ["TESTING"] = "1"

import pytest
import tempfile
import shutil
import json
from typing import Generator, Dict, Any, List
from unittest.mock import Mock, MagicMock, patch
from io import BytesIO
import sys
import socket
import subprocess
import time

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import httpx
from fastapi.testclient import TestClient
from src.api.main import app
from src.services.masterfile_service import MasterfileService
from src.services.json_service import JsonService
from src.services.dbn_service import DBNService
from src.api.dependencies import TempFileManager, JobRegistry


@pytest.fixture(scope="module")
def client():
    """TestClient with module scope to avoid event loop issues across tests"""
    with TestClient(app, raise_server_exceptions=False) as test_client:
        yield test_client


def _get_free_tcp_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return int(s.getsockname()[1])


@pytest.fixture(scope="session")
def api_server() -> str:
    """Starts a real uvicorn server for tests that depend on BackgroundTasks."""
    port = _get_free_tcp_port()
    base_url = f"http://127.0.0.1:{port}"

    # Use python -m uvicorn to ensure we use the same interpreter/env.
    cmd = [
        sys.executable,
        "-m",
        "uvicorn",
        "src.api.main:app",
        "--host",
        "127.0.0.1",
        "--port",
        str(port),
        "--log-level",
        "warning",
    ]

    proc = subprocess.Popen(
        cmd,
        cwd=os.path.join(os.path.dirname(__file__), ".."),
        env={**os.environ, "PYTHONUNBUFFERED": "1"},
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
    )

    # Wait until server is ready.
    deadline = time.time() + 30
    last_error: Exception | None = None
    while time.time() < deadline:
        try:
            r = httpx.get(f"{base_url}/health", timeout=1.0)
            if r.status_code == 200:
                last_error = None
                break
        except Exception as e:
            last_error = e
        time.sleep(0.2)

    if last_error is not None:
        try:
            proc.terminate()
        except Exception:
            pass
        raise RuntimeError(f"API server did not start in time: {last_error}")

    yield base_url

    try:
        proc.terminate()
        proc.wait(timeout=10)
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


@pytest.fixture
def api_client(api_server: str):
    """httpx client pointing to the real running API server."""
    with httpx.Client(base_url=api_server, timeout=120.0) as c:
        yield c


@pytest.fixture
def temp_directory():
    temp_dir = tempfile.mkdtemp(prefix="test_automation_")
    yield temp_dir
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir, ignore_errors=True)


@pytest.fixture
def output_directory(temp_directory):
    output_dir = os.path.join(temp_directory, "output")
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


@pytest.fixture
def sample_excel_file(temp_directory):
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        headers = ["Inventory Name", "Column A", "Column B"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        data = [["INV_TEST_A", "V1", "V2"], ["INV_TEST_B", "V3", "V4"]]
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        file_path = os.path.join(temp_directory, "test_pack.xlsx")
        wb.save(file_path)
        return file_path
    except ImportError:
        file_path = os.path.join(temp_directory, "test_pack.xlsx")
        with open(file_path, 'wb') as f:
            f.write(b'')
        return file_path


@pytest.fixture
def sample_excel_bytes():
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Inventory Name")
        ws.cell(row=2, column=1, value="INV_TEST_A")
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except ImportError:
        return b'PK\x03\x04'


@pytest.fixture
def sample_excel_path(temp_directory):
    return os.path.join(temp_directory, "test_pack.xlsx")


@pytest.fixture
def sample_inventory_names():
    return ["INV_TEST_A", "INV_TEST_B"]


@pytest.fixture
def sample_dbn_names():
    return ["DBN_TEST_001", "DBN_TEST_002"]


@pytest.fixture
def valid_db_types():
    return ["oracle", "postgres", "postgresql"]


@pytest.fixture
def masterfile_service():
    return MasterfileService()


@pytest.fixture
def json_service():
    return JsonService()


@pytest.fixture
def dbn_service():
    return DBNService()


@pytest.fixture
def mock_masterfile_creator():
    with patch('src.services.masterfile_service.MasterfileCreator') as mock:
        mock_instance = MagicMock()
        mock_instance.create_masterfiles.return_value = True
        mock.return_value = mock_instance
        yield mock


@pytest.fixture
def mock_json_creator():
    with patch('src.services.json_service.JsonCreator') as mock:
        mock_instance = MagicMock()
        mock_instance.create_json.return_value = True
        mock.return_value = mock_instance
        yield mock


@pytest.fixture
def mock_dbn_model():
    with patch('src.services.dbn_service.DBNModel') as mock:
        mock.modelo_exportacao.return_value = True
        yield mock


@pytest.fixture
def mock_file_repository():
    mock = MagicMock()
    mock.list_files.return_value = ["file1.json", "file2.json"]
    return mock


@pytest.fixture
def temp_file_manager():
    manager = TempFileManager()
    yield manager
    manager.cleanup()


@pytest.fixture
def job_registry():
    JobRegistry._jobs.clear()
    return JobRegistry()


@pytest.fixture
def api_form_data(sample_inventory_names):
    return {
        "inventory_names": json.dumps(sample_inventory_names),
        "db_type": "oracle"
    }


@pytest.fixture
def dbn_form_data(sample_dbn_names, output_directory):
    return {
        "dbn_names": json.dumps(sample_dbn_names),
        "output_path": output_directory,
        "schema": "TEST_SCHEMA"
    }
